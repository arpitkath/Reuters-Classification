from textblob.classifiers import NaiveBayesClassifier
import openpyxl
import preprocess
from random import randint
from nltk.corpus import stopwords
from collections import Counter
import re

wb = openpyxl.load_workbook('input.xlsx')

training_sheet = wb.get_sheet_by_name("training")
data_sheet = wb.get_sheet_by_name("data")

training_set = []
#test_set = []
count = 0


def tokenize(body):
    #body = re.sub("[^a-zA-Z<>/\s=!-\"\"]+","", body)
    #body = [word for word in body.split() if word not in stopwords.words('english')]
    return dict(Counter(body))

for i in range(2, training_sheet.max_row+1):
    topic = training_sheet.cell(row=i, column=1).value
    body = training_sheet.cell(row=i, column=2).value
    if topic is None or body is None:
        continue
    tag = training_sheet.cell(row=i, column=1).value.split(",")
    tag = tag[randint(0, len(tag)-1)]
    body = tokenize(body)
    tup = (body, tag)
    count += 1
    training_set.append(tup)
print("Preparing training set.")

cl = NaiveBayesClassifier(training_set)
#print("Accuracy of the classifier: {0}".format(cl.accuracy(test_set)))
#cl.show_informative_features(10)
print("Getting result.")
for i in range(3, data_sheet.max_row):
    data_sheet.cell(row=i, column=1).value = cl.classify(tokenize(data_sheet.cell(row=i, column=3).value))
wb.save("output.xlsx")