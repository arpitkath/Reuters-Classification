from bs4 import BeautifulSoup
import openpyxl
import re
from nltk.corpus import stopwords

wb = openpyxl.Workbook()
wb.create_sheet(index=0, title="training")
wb.create_sheet(index=1, title="data")
training_sheet = wb.get_sheet_by_name("training")
data_sheet = wb.get_sheet_by_name("data")
tags = ["TOPICS", "BODY"]
training_count = [0]*len(tags)
d_count = 0
for i in range(1, len(tags)+1):
    training_sheet.cell(row=1, column=i).value = tags[i-1]
    data_sheet.cell(row=1, column=i).value = tags[i-1]

stop = set(stopwords.words('english'))

def filter_body(body):
    body = re.sub("[^a-zA-Z<>/\s=!-\"\"]+","", body)
    body = [word for word in body.split() if word not in stop]
    return " ".join(str(i) for i in body)

def parse(file_name):
    f = open(file_name, 'r')
    data = f.read()
    soup = BeautifulSoup(data, 'html.parser')
    global training_count, d_count
    reuters = soup.find_all("reuters")
    for i in range(len(reuters)):# Gettinng if it has topic list or not.
        reuters[i] = reuters[i]["topics"]
    for tag in range(1, len(tags)+1):
        content = soup.find_all(tags[tag-1].lower())
        t_count = training_count[tag-1]
        #print(len(content))
        for i in range(len(content)):
            if reuters[i] == "YES":
                if len(content[i]) == 1:
                    try:
                        training_sheet.cell(row=t_count+2, column=tag).value = filter_body(content[i].text)
                    except:
                        continue
                else:
                    s = ",".join(str(j.text) for j in content[i])
                    training_sheet.cell(row=t_count+2, column=tag).value = s
                t_count += 1
        training_count[tag-1] = t_count
    content = soup.find_all(tags[-1].lower())
    for i in range(len(reuters)):
        if reuters[i] == "NO":
            try:
                data_sheet.cell(row=d_count+2, column=len(tags)+1).value = filter_body(content[i].text)
                d_count += 1
            except:
                continue

number_of_files = 22

def make_xls():
    for i in range(number_of_files):
        if i < 10:
            file_name = "reut2-00"+str(i)+".sgm"
        else:
            file_name = "reut2-0"+str(i)+".sgm"
        parse(file_name)
        print(file_name)
make_xls()
wb.save("input.xlsx")
